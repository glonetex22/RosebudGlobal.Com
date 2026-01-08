from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "New Items"

# Styles
header_font = Font(bold=True, color='FFFFFF', size=12)
header_fill = PatternFill('solid', fgColor='D63585')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ['Model #', 'Product Name', 'Description', 'Category', 'Price', 'Status']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# New Items Data
items = [
    ['RBG-C2600', 'Custom 100% Cotton Dyed Terry Hand Towels', 'Wholesale custom dyed terry hand towels, 100% cotton', 'Custom Gift Items', 'Contact For Pricing', 'NEW'],
    ['RBG-P31005', 'Special Shaped Plate Series Latest Design Ceramic Plate', 'Wholesale ceramic plates with special shaped designs', 'Fine China', 'Contact For Pricing', 'NEW'],
    ['RBG-F29055', 'Custom Exotic Material Loveseat', 'Custom designed loveseat with exotic material upholstery', 'Furniture', 'Contact For Pricing', 'NEW'],
    ['RT1189', 'Olivia Riegel Priscilla Picture Frame', 'Elegant crystal picture frame by Olivia Riegel', 'Home Decor', '$300.00', 'NEW'],
]

for row_num, item in enumerate(items, 2):
    for col_num, value in enumerate(item, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.border = border
        if col_num == 6:  # Status column
            cell.font = Font(color='38CB89', bold=True)

# Column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 10

wb.save('inventory/new_items_inventory.xlsx')
print("Inventory spreadsheet created")
